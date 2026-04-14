import streamlit as st
import cv2
import numpy as np
from PIL import Image
import io
import re
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import datetime
 
st.set_page_config(page_title="ReceiptIQ · FIN4901", page_icon="🧾", layout="wide", initial_sidebar_state="collapsed")
 
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Syne:wght@400;600;700;800&family=DM+Mono:wght@300;400;500&display=swap');
:root{--bg:#07090e;--s1:#0c0f17;--s2:#11151f;--border:#1a2030;--accent:#00e5b0;--accent2:#6c63ff;--warn:#f5a623;--text:#d8e0f0;--muted:#4e5a78;--r:10px}
html,body,[data-testid="stAppViewContainer"]{background:var(--bg)!important;color:var(--text)!important;font-family:'DM Mono',monospace}
[data-testid="stSidebar"]{background:var(--s1)!important}
#MainMenu,footer,header{visibility:hidden}
[data-testid="stDecoration"]{display:none}
h1,h2,h3,h4{font-family:'Syne',sans-serif!important}
.hero{padding:3rem 3.5rem 2.5rem;background:linear-gradient(160deg,#0b1120 0%,#090d18 55%,#060810 100%);border:1px solid var(--border);border-radius:18px;margin-bottom:2rem;position:relative;overflow:hidden}
.hero::after{content:'';position:absolute;width:700px;height:700px;background:radial-gradient(circle,rgba(0,229,176,.055) 0%,transparent 65%);top:-260px;right:-120px;pointer-events:none}
.hero-eyebrow{font-size:.68rem;letter-spacing:.18em;text-transform:uppercase;color:var(--accent);margin-bottom:.55rem;font-family:'DM Mono',monospace}
.hero-title{font-family:'Syne',sans-serif;font-size:clamp(2rem,5vw,3.5rem);font-weight:800;line-height:1.05;margin:0 0 .7rem;background:linear-gradient(100deg,#00e5b0 0%,#6c63ff 100%);-webkit-background-clip:text;-webkit-text-fill-color:transparent;background-clip:text}
.hero-desc{color:var(--muted);font-size:.88rem;margin:0;line-height:1.75}
.badge{display:inline-flex;align-items:center;gap:.4rem;font-size:.68rem;letter-spacing:.08em;text-transform:uppercase;padding:.28rem .85rem;border-radius:100px;font-family:'DM Mono',monospace;margin-right:.5rem}
.badge-green{background:rgba(0,229,176,.1);border:1px solid rgba(0,229,176,.25);color:var(--accent)}
.badge-purple{background:rgba(108,99,255,.1);border:1px solid rgba(108,99,255,.25);color:#a09fff}
.badge-warn{background:rgba(245,166,35,.1);border:1px solid rgba(245,166,35,.25);color:var(--warn)}
.stats{display:flex;gap:.85rem;flex-wrap:wrap;margin:1.4rem 0}
.stat{flex:1;min-width:105px;background:var(--s2);border:1px solid var(--border);border-radius:var(--r);padding:.9rem 1.1rem;text-align:center}
.stat-val{font-family:'Syne',sans-serif;font-size:1.6rem;font-weight:700;color:var(--accent);display:block;line-height:1.1}
.stat-lbl{font-size:.62rem;color:var(--muted);text-transform:uppercase;letter-spacing:.1em}
.section-hdr{font-family:'Syne',sans-serif;font-size:1.1rem;font-weight:700;color:var(--text);margin:1.8rem 0 .9rem}
.rcpt-card{background:var(--s1);border:1px solid var(--border);border-radius:14px;margin:1.25rem 0;overflow:hidden}
.rcpt-header{background:linear-gradient(90deg,rgba(0,229,176,.07),rgba(108,99,255,.04));border-bottom:1px solid var(--border);padding:1rem 1.5rem;display:flex;justify-content:space-between;align-items:flex-start}
.rcpt-num{font-size:.63rem;color:var(--accent);text-transform:uppercase;letter-spacing:.12em;font-family:'DM Mono',monospace}
.rcpt-store{font-family:'Syne',sans-serif;font-size:1.15rem;font-weight:700;color:var(--text);margin:.15rem 0}
.rcpt-date{font-size:.78rem;color:var(--muted)}
.rcpt-total{font-family:'Syne',sans-serif;font-size:1.35rem;font-weight:700;color:var(--accent);white-space:nowrap}
.rcpt-body{padding:1.1rem 1.5rem}
.itbl{width:100%;border-collapse:collapse;font-size:.81rem}
.itbl th{background:var(--s2);color:var(--muted);font-size:.6rem;text-transform:uppercase;letter-spacing:.1em;padding:.45rem .7rem;text-align:left;font-family:'Syne',sans-serif}
.itbl td{padding:.42rem .7rem;border-bottom:1px solid var(--border);color:var(--text)}
.itbl tr:last-child td{border-bottom:none}
.itbl .pr{text-align:right;color:var(--accent);font-weight:500}
.itbl .tot-row td{color:var(--accent);font-weight:700;border-top:1px solid rgba(0,229,176,.25);font-family:'Syne',sans-serif}
.tip{background:rgba(108,99,255,.06);border:1px solid rgba(108,99,255,.16);border-radius:var(--r);padding:.85rem 1.05rem;font-size:.79rem;color:var(--muted);line-height:1.65;margin:.65rem 0}
.tip strong{color:#a09fff}
.edit-box{background:rgba(0,229,176,.04);border:1px solid rgba(0,229,176,.18);border-radius:var(--r);padding:1rem 1.25rem;margin:.75rem 0}
.analysis-card{background:var(--s1);border:1px solid var(--border);border-radius:14px;padding:1.25rem 1.5rem;margin:1rem 0}
.best-row{display:flex;justify-content:space-between;align-items:center;background:var(--s2);border:1px solid var(--border);border-radius:8px;padding:.65rem .95rem;margin:.3rem 0;font-size:.82rem}
.best-name{color:var(--text)}.best-store{color:var(--muted);font-size:.72rem}.best-price{color:var(--accent);font-weight:600}
.best-tag{font-size:.6rem;background:rgba(0,229,176,.12);color:var(--accent);border-radius:4px;padding:2px 7px;margin-left:5px;text-transform:uppercase;letter-spacing:.06em}
.bar-wrap{margin:.5rem 0}.bar-row{display:flex;align-items:center;gap:.75rem;margin:.35rem 0;font-size:.8rem}
.bar-label{color:var(--muted);width:130px;flex-shrink:0;overflow:hidden;text-overflow:ellipsis;white-space:nowrap}
.bar-track{flex:1;background:var(--s2);border-radius:100px;height:8px;overflow:hidden}
.bar-fill{height:100%;border-radius:100px;background:linear-gradient(90deg,var(--accent),var(--accent2))}
.bar-value{color:var(--accent);font-weight:600;width:60px;text-align:right;flex-shrink:0}
.stButton>button{background:linear-gradient(135deg,#00e5b0,#00c497)!important;color:#031a13!important;font-family:'Syne',sans-serif!important;font-weight:700!important;font-size:.87rem!important;border:none!important;border-radius:9px!important;height:auto!important;padding:.68rem 1.4rem!important;width:100%}
.stButton>button:hover{opacity:.85!important}
[data-testid="stDownloadButton"]>button{background:var(--s2)!important;color:var(--accent)!important;border:1px solid rgba(0,229,176,.28)!important;font-family:'Syne',sans-serif!important;font-weight:600!important;border-radius:9px!important;height:auto!important;width:100%}
[data-testid="stDownloadButton"]>button:hover{border-color:var(--accent)!important}
[data-baseweb="tab-list"]{background:var(--s2)!important;border-radius:9px!important;padding:4px!important;border:1px solid var(--border)!important;gap:3px!important}
[data-baseweb="tab"]{background:transparent!important;border-radius:7px!important;color:var(--muted)!important;font-family:'Syne',sans-serif!important;font-weight:600!important;font-size:.82rem!important}
[aria-selected="true"][data-baseweb="tab"]{background:var(--s1)!important;color:var(--accent)!important}
[data-testid="stFileUploader"]{background:var(--s2)!important;border:2px dashed var(--border)!important;border-radius:var(--r)!important}
[data-testid="stFileUploader"]:hover{border-color:var(--accent)!important}
[data-testid="stCameraInput"]>div{background:var(--s2)!important;border:2px solid var(--border)!important;border-radius:var(--r)!important}
.stProgress>div>div{background:linear-gradient(90deg,var(--accent),var(--accent2))!important;border-radius:100px!important}
[data-testid="stExpander"]{background:var(--s2)!important;border:1px solid var(--border)!important;border-radius:var(--r)!important}
button:focus-visible,input:focus-visible,a:focus-visible{outline:2px solid var(--accent)!important;outline-offset:3px!important}
hr{border-color:var(--border)!important}
/* Style Streamlit text inputs inside edit boxes to match dark theme */
.stTextInput>div>div>input{background:var(--s2)!important;color:var(--text)!important;border:1px solid var(--border)!important;border-radius:6px!important;font-family:'DM Mono',monospace!important}
.stNumberInput>div>div>input{background:var(--s2)!important;color:var(--text)!important;border:1px solid var(--border)!important;border-radius:6px!important;font-family:'DM Mono',monospace!important}
@media(max-width:768px){.hero{padding:1.4rem}.hero-title{font-size:2rem}.stats{gap:.5rem}.rcpt-header{flex-direction:column;gap:.5rem}}
</style>
""", unsafe_allow_html=True)
 
# ── Session state ─────────────────────────────────────────────
if "receipts" not in st.session_state:
    st.session_state["receipts"] = []
 
# ── Model loading ─────────────────────────────────────────────
@st.cache_resource(show_spinner=False)
def load_ocr():
    import easyocr
    return easyocr.Reader(['en'], gpu=False, verbose=False)
 
@st.cache_resource(show_spinner=False)
def load_engine():
    try:
        from yolo_engine import ReceiptEngine
        engine = ReceiptEngine(conf=0.30, buffer=20)
        return engine, f"YOLO26 trained · {engine.model_path}"
    except FileNotFoundError:
        return None, "contour fallback"
    except Exception as e:
        return None, f"contour fallback ({e})"
 
# ── Contour fallback ──────────────────────────────────────────
def detect_contours(img_pil):
    img=np.array(img_pil); gray=cv2.cvtColor(img,cv2.COLOR_RGB2GRAY)
    edge=cv2.Canny(cv2.GaussianBlur(gray,(5,5),0),30,100)
    dil=cv2.dilate(edge,np.ones((15,15),np.uint8),iterations=2)
    cnts,_=cv2.findContours(dil,cv2.RETR_EXTERNAL,cv2.CHAIN_APPROX_SIMPLE)
    H,W=img.shape[:2]; crops,boxes=[],[]
    for c in cnts:
        if cv2.contourArea(c)<H*W*0.04: continue
        x,y,bw,bh=cv2.boundingRect(c)
        if not (0.4<bh/max(bw,1)<9): continue
        pad=12; x1,y1=max(0,x-pad),max(0,y-pad); x2,y2=min(W,x+bw+pad),min(H,y+bh+pad)
        dup=any(max(0,min(x2,bx2)-max(x1,bx1))*max(0,min(y2,by2)-max(y1,by1))/max((x2-x1)*(y2-y1)+(bx2-bx1)*(by2-by1)-max(0,min(x2,bx2)-max(x1,bx1))*max(0,min(y2,by2)-max(y1,by1)),1)>0.7 for (bx1,by1,bx2,by2) in boxes)
        if not dup: crops.append(img_pil.crop((x1,y1,x2,y2))); boxes.append((x1,y1,x2,y2))
    return crops if crops else [img_pil]
 
def detect_receipts(img_pil,engine,status):
    if engine is not None:
        crops=engine.detect_and_crop_all(img_pil,return_pil=True)
        if crops: return crops,status
    return detect_contours(img_pil),"contour fallback"
 
# ── OCR + parsing ─────────────────────────────────────────────
def run_ocr(img_pil,reader):
    return [(t.strip(),float(c)) for (_,t,c) in reader.readtext(np.array(img_pil)) if c>0.35]
 
def parse_receipt(lines):
    texts=[t for t,_ in lines]
    store=texts[0].title() if texts else "Unknown Store"
    date=None; items=[]; total=subtotal=tax=None
    for line in texts:
        for pat in [r'\b(\d{1,2}[\/\-\.]\d{1,2}[\/\-\.]\d{2,4})\b',r'\b(\d{4}[\/\-\.]\d{1,2}[\/\-\.]\d{1,2})\b',r'\b(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)[a-z]*\.?\s+\d{1,2},?\s+\d{4}\b']:
            m=re.search(pat,line,re.IGNORECASE)
            if m: date=m.group(0); break
        if date: break
    price_re=re.compile(r'\$?\s*(\d+\.\d{2})')
    skip={'total','subtotal','sub total','sub-total','tax','change','cash','card','tip','visa','mastercard','amex','balance','due'}
    for line in texts:
        low=line.lower(); prices=price_re.findall(line)
        if not prices: continue
        amt=float(prices[-1])
        if any(k in low for k in ['total','balance','due']):
            if total is None or amt>total: total=amt
        elif any(k in low for k in ['subtotal','sub total','sub-total']): subtotal=amt
        elif 'tax' in low: tax=amt
        elif not any(k in low for k in skip):
            name=re.sub(r'\s{2,}',' ',price_re.sub('',line)).strip().strip('$-').strip()
            if name and len(name)>1: items.append({"name":name.title(),"price":amt})
    return {"store_name":store,"date":date or "Unknown","items":items,"subtotal":subtotal,"tax":tax,"total":total,"raw_lines":texts}
 
# ── Excel export ──────────────────────────────────────────────
def build_excel(receipts):
    wb=openpyxl.Workbook()
    BG="07090e";HDR="00e5b0";DIM="4e5a78";ALT="0c0f17";TXT="d8e0f0";TOT="00e5b0";ACC="0b1120"
    def hc(ws,r,c,v,w=None):
        cell=ws.cell(row=r,column=c,value=v); cell.font=Font(bold=True,color=HDR,name="Courier New",size=9)
        cell.fill=PatternFill("solid",fgColor=ACC); cell.alignment=Alignment(horizontal="center",vertical="center")
        if w: ws.column_dimensions[get_column_letter(c)].width=w
    def dc(ws,r,c,v,bold=False,color=TXT,fmt=None,align="left",bg=None):
        cell=ws.cell(row=r,column=c,value=v); cell.font=Font(bold=bold,color=color,name="Courier New",size=9)
        cell.alignment=Alignment(horizontal=align,vertical="center")
        if fmt: cell.number_format=fmt
        if bg: cell.fill=PatternFill("solid",fgColor=bg)
    ws=wb.active; ws.title="Summary"; ws.sheet_view.showGridLines=False
    ws.row_dimensions[1].height=32; ws.merge_cells("A1:G1")
    ws["A1"].value="RECEIPTIQ  ·  SUMMARY REPORT"; ws["A1"].font=Font(bold=True,color=HDR,name="Courier New",size=13)
    ws["A1"].fill=PatternFill("solid",fgColor=ACC); ws["A1"].alignment=Alignment(horizontal="left",vertical="center")
    ws.merge_cells("A2:G2"); ws["A2"].value=f"Generated: {datetime.datetime.now().strftime('%Y-%m-%d  %H:%M:%S')}"
    ws["A2"].font=Font(color=DIM,name="Courier New",size=8); ws["A2"].fill=PatternFill("solid",fgColor=BG)
    for i,(h,w) in enumerate(zip(["#","Store","Date","Items","Subtotal","Tax","Total"],[4,32,18,8,14,14,14]),1): hc(ws,4,i,h,w)
    grand=0.0
    for i,rd in enumerate(receipts):
        r=5+i; bg=ALT if i%2 else BG
        dc(ws,r,1,i+1,align="center",bg=bg,color=DIM); dc(ws,r,2,rd["store_name"],bg=bg)
        dc(ws,r,3,rd["date"],bg=bg,color=DIM); dc(ws,r,4,len(rd["items"]),align="center",bg=bg,color=DIM)
        dc(ws,r,5,rd["subtotal"] or "",fmt='$#,##0.00',align="right",bg=bg,color=DIM)
        dc(ws,r,6,rd["tax"] or "",fmt='$#,##0.00',align="right",bg=bg,color=DIM)
        tot=rd["total"] or 0; dc(ws,r,7,tot,fmt='$#,##0.00',align="right",bg=bg,bold=True,color=TOT)
        if isinstance(tot,(int,float)): grand+=tot
    tr=5+len(receipts); ws.row_dimensions[tr].height=20; ws.merge_cells(f"A{tr}:F{tr}")
    dc(ws,tr,1,"GRAND TOTAL",bold=True,color=HDR,bg="001a14",align="right")
    dc(ws,tr,7,grand,fmt='$#,##0.00',align="right",bold=True,color=HDR,bg="001a14")
    ws2=wb.create_sheet("Line Items"); ws2.sheet_view.showGridLines=False
    ws2.merge_cells("A1:E1"); ws2["A1"].value="ALL LINE ITEMS"
    ws2["A1"].font=Font(bold=True,color=HDR,name="Courier New",size=11); ws2["A1"].fill=PatternFill("solid",fgColor=ACC)
    for i,(h,w) in enumerate(zip(["Receipt #","Store","Date","Item","Price"],[10,30,16,44,14]),1): hc(ws2,3,i,h,w)
    row=4
    for i,rd in enumerate(receipts):
        bg=ALT if i%2 else BG
        for item in rd["items"]:
            dc(ws2,row,1,i+1,align="center",bg=bg,color=DIM); dc(ws2,row,2,rd["store_name"],bg=bg)
            dc(ws2,row,3,rd["date"],bg=bg,color=DIM); dc(ws2,row,4,item["name"],bg=bg)
            dc(ws2,row,5,item["price"],fmt='$#,##0.00',align="right",bold=True,color=TOT,bg=bg); row+=1
    ws3=wb.create_sheet("Price Comparison"); ws3.sheet_view.showGridLines=False
    ws3.merge_cells("A1:D1"); ws3["A1"].value="PRICE COMPARISON  ·  BEST DEALS"
    ws3["A1"].font=Font(bold=True,color=HDR,name="Courier New",size=11); ws3["A1"].fill=PatternFill("solid",fgColor=ACC)
    for i,(h,w) in enumerate(zip(["Item","Store","Date","Price"],[44,30,16,14]),1): hc(ws3,3,i,h,w)
    item_map={}
    for rd in receipts:
        for it in rd["items"]: item_map.setdefault(it["name"].lower().strip(),[]).append({"store":rd["store_name"],"date":rd["date"],"price":it["price"]})
    dupes={k:v for k,v in item_map.items() if len(v)>1}
    if dupes:
        row=4
        for name,entries in dupes.items():
            best=min(e["price"] for e in entries)
            for e in sorted(entries,key=lambda x:x["price"]):
                ib=e["price"]==best; bg="001a14" if ib else BG; col=TOT if ib else TXT
                dc(ws3,row,1,name.title()+(" ★ BEST" if ib else ""),bold=ib,color=col,bg=bg)
                dc(ws3,row,2,e["store"],color=col,bg=bg); dc(ws3,row,3,e["date"],color=DIM,bg=bg)
                dc(ws3,row,4,e["price"],fmt='$#,##0.00',align="right",bold=ib,color=col,bg=bg); row+=1
    else:
        ws3["A4"].value="No duplicate items found — scan receipts from multiple stores to compare."
        ws3["A4"].font=Font(color=DIM,name="Courier New",size=9)
    buf=io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf.getvalue()
 
# ── Main pipeline ─────────────────────────────────────────────
def process_image(source,reader,engine,engine_status):
    source.seek(0)
    arr=np.frombuffer(source.read(),dtype=np.uint8)
    img_cv=cv2.imdecode(arr,cv2.IMREAD_COLOR)
    if img_cv is None:
        st.error("Could not decode image — please use JPG or PNG."); return []
    img_pil=Image.fromarray(cv2.cvtColor(img_cv,cv2.COLOR_BGR2RGB))
    with st.spinner("Detecting receipt regions…"):
        crops,det_label=detect_receipts(img_pil,engine,engine_status)
    short_det=det_label.split("·")[0].strip()
    st.markdown(f"""<div class="stats">
      <div class="stat"><span class="stat-val">{len(crops)}</span><span class="stat-lbl">Receipts found</span></div>
      <div class="stat"><span class="stat-val" style="font-size:.9rem;padding-top:.35rem;">{short_det}</span><span class="stat-lbl">Detector</span></div>
      <div class="stat"><span class="stat-val">{img_pil.width}×{img_pil.height}</span><span class="stat-lbl">Resolution</span></div>
    </div>""",unsafe_allow_html=True)
    if engine is not None and "trained" in det_label:
        with st.expander("🔍 Detection overlay",expanded=False):
            st.image(engine.annotate(img_pil),use_column_width=True)
    parsed_list=[]; prog=st.progress(0,text="Starting OCR…")
    for i,crop in enumerate(crops):
        prog.progress(i/len(crops),text=f"Reading receipt {i+1} of {len(crops)}…")
        parsed=parse_receipt(run_ocr(crop,reader))
        parsed["crop"]=crop
        parsed["_idx"]=len(st.session_state["receipts"])+len(parsed_list)
        parsed_list.append(parsed)
        store=parsed["store_name"]; date=parsed["date"]
        total=parsed["total"]; items=parsed["items"]
        tot_html=(f'<div class="rcpt-total">${total:.2f}</div>' if isinstance(total,float) else "")
        st.markdown(f"""<div class="rcpt-card">
          <div class="rcpt-header">
            <div><div class="rcpt-num">Receipt {i+1} of {len(crops)}</div>
            <div class="rcpt-store">{store}</div>
            <div class="rcpt-date">📅 &nbsp;{date}</div></div>{tot_html}
          </div><div class="rcpt-body">""",unsafe_allow_html=True)
        c1,c2=st.columns([1,2],gap="large")
        with c1: st.image(crop,use_column_width=True)
        with c2:
            if items:
                rows="".join(f'<tr><td>{it["name"]}</td><td class="pr">${it["price"]:.2f}</td></tr>' for it in items)
                tot_row=(f'<tr class="tot-row"><td>Total</td><td class="pr">${total:.2f}</td></tr>' if isinstance(total,float) else "")
                st.markdown(f'<table class="itbl"><thead><tr><th>Item</th><th style="text-align:right">Price</th></tr></thead><tbody>{rows}{tot_row}</tbody></table>',unsafe_allow_html=True)
            else:
                st.markdown('<div class="tip">⚠️ No line items detected — try better lighting or move closer.</div>',unsafe_allow_html=True)
            buf=io.BytesIO(); crop.save(buf,format="JPEG",quality=93)
            cs=re.sub(r'[^a-zA-Z0-9]','',store); cd=re.sub(r'[^0-9\-]','',date).replace("/","-")
            st.download_button(label="💾 Save receipt image",data=buf.getvalue(),
                file_name=f"{cd}_{cs}_{i+1}.jpg",mime="image/jpeg",
                key=f"img_{i}_{len(st.session_state['receipts'])}")
        st.markdown("</div></div>",unsafe_allow_html=True)
        with st.expander(f"Raw OCR text — receipt {i+1}",expanded=False):
            st.code("\n".join(parsed["raw_lines"]) or "(none)",language=None)
    prog.progress(1.0,text="✅ Done!")
    return parsed_list
 
# ── Manual edit panel ─────────────────────────────────────────
def render_edit_panel(all_r):
    st.markdown('<div class="section-hdr">✏️ Review & Correct Data</div>',unsafe_allow_html=True)
    st.markdown('<div class="tip"><strong>Check the extracted data below.</strong> If anything is wrong — store name, date, prices — edit it here before exporting. Changes apply immediately to the Excel report and file names.</div>',unsafe_allow_html=True)
 
    for i, rd in enumerate(all_r):
        with st.expander(f"Receipt {i+1} — {rd['store_name']}  ·  {rd['date']}", expanded=False):
            col_img, col_fields = st.columns([1, 2], gap="large")
 
            with col_img:
                if "crop" in rd:
                    st.image(rd["crop"], use_column_width=True)
 
            with col_fields:
                st.markdown("**Store name & date**")
                new_store = st.text_input("Store name", value=rd["store_name"], key=f"edit_store_{i}")
                new_date  = st.text_input("Date", value=rd["date"], key=f"edit_date_{i}")
 
                st.markdown("**Totals**")
                c1, c2, c3 = st.columns(3)
                with c1:
                    sub_val = rd["subtotal"] if isinstance(rd.get("subtotal"), float) else 0.0
                    new_sub = st.number_input("Subtotal", value=sub_val, min_value=0.0, step=0.01, format="%.2f", key=f"edit_sub_{i}")
                with c2:
                    tax_val = rd["tax"] if isinstance(rd.get("tax"), float) else 0.0
                    new_tax = st.number_input("Tax", value=tax_val, min_value=0.0, step=0.01, format="%.2f", key=f"edit_tax_{i}")
                with c3:
                    tot_val = rd["total"] if isinstance(rd.get("total"), float) else 0.0
                    new_tot = st.number_input("Total", value=tot_val, min_value=0.0, step=0.01, format="%.2f", key=f"edit_tot_{i}")
 
                st.markdown("**Line items**")
                new_items = []
                for j, item in enumerate(rd["items"]):
                    ic1, ic2 = st.columns([3, 1])
                    with ic1:
                        new_name = st.text_input("Item", value=item["name"], key=f"edit_item_name_{i}_{j}", label_visibility="collapsed")
                    with ic2:
                        new_price = st.number_input("Price", value=item["price"], min_value=0.0, step=0.01, format="%.2f", key=f"edit_item_price_{i}_{j}", label_visibility="collapsed")
                    new_items.append({"name": new_name, "price": new_price})
 
                # Apply changes back to session state
                st.session_state["receipts"][i]["store_name"] = new_store
                st.session_state["receipts"][i]["date"]       = new_date
                st.session_state["receipts"][i]["subtotal"]   = new_sub if new_sub > 0 else None
                st.session_state["receipts"][i]["tax"]        = new_tax if new_tax > 0 else None
                st.session_state["receipts"][i]["total"]      = new_tot if new_tot > 0 else None
                st.session_state["receipts"][i]["items"]      = new_items
 
                # Updated download button using corrected name/date
                if "crop" in rd:
                    buf = io.BytesIO()
                    rd["crop"].save(buf, format="JPEG", quality=93)
                    cs = re.sub(r'[^a-zA-Z0-9]', '', new_store)
                    cd = re.sub(r'[^0-9\-]', '', new_date).replace("/", "-")
                    st.download_button(
                        label=f"💾 Save as  {cs}_{cd}.jpg",
                        data=buf.getvalue(),
                        file_name=f"{cs}_{cd}.jpg",
                        mime="image/jpeg",
                        key=f"edit_dl_{i}"
                    )
 
# ── Header ────────────────────────────────────────────────────
st.markdown("""<div class="hero" role="banner">
  <div class="hero-eyebrow">FIN4901 · FinTech Graduate Project</div>
  <div class="hero-title">ReceiptIQ</div>
  <p class="hero-desc">Multi-receipt detection &nbsp;·&nbsp; YOLO26 trained model &nbsp;·&nbsp; EasyOCR extraction &nbsp;·&nbsp; Manual correction &nbsp;·&nbsp; Excel export</p>
</div>""",unsafe_allow_html=True)
 
# ── Load models ───────────────────────────────────────────────
col_ocr,col_yolo=st.columns(2)
with col_ocr:
    with st.spinner("Loading EasyOCR…"): reader=load_ocr()
    if reader: st.markdown('<span class="badge badge-green">● EasyOCR ready</span>',unsafe_allow_html=True)
    else: st.error("EasyOCR failed to load"); st.stop()
with col_yolo:
    with st.spinner("Loading YOLO26…"): engine,engine_status=load_engine()
    if engine is not None:
        st.markdown(f'<span class="badge badge-green">● {engine_status}</span>',unsafe_allow_html=True)
    else:
        st.markdown('<span class="badge badge-warn">◐ Trained weights not found — using contour fallback</span>',unsafe_allow_html=True)
        with st.expander("📋 How to deploy your trained YOLO26 weights",expanded=False):
            st.markdown("""
1. Rename `best.pt` → `receipt_detector.pt`
2. Place it in the repo root alongside `streamlit_app.py`
3. Commit & push — the green badge confirms YOLO26 is active
            """)
 
# ── Input tabs ────────────────────────────────────────────────
tab_cam,tab_upload=st.tabs(["📷   Camera","📁   Upload"])
with tab_cam:
    st.markdown('<div class="tip"><strong>Tips:</strong> flat dark surface · even lighting · steady hold · one or multiple receipts per frame</div>',unsafe_allow_html=True)
    cam=st.camera_input("Capture",label_visibility="collapsed")
    if cam:
        res=process_image(cam,reader,engine,engine_status)
        if res: st.session_state["receipts"]+=res
with tab_upload:
    st.markdown('<div class="tip"><strong>Supported:</strong> JPG, JPEG, PNG — one image can contain multiple receipts</div>',unsafe_allow_html=True)
    up=st.file_uploader("Drop image here",type=["jpg","jpeg","png"],label_visibility="collapsed")
    if up:
        res=process_image(up,reader,engine,engine_status)
        if res: st.session_state["receipts"]+=res
 
# ── Review, analytics + export ────────────────────────────────
all_r=st.session_state["receipts"]
if all_r:
    st.markdown("---")
 
    # Stats
    total_items=sum(len(r["items"]) for r in all_r)
    grand=sum(r["total"] for r in all_r if isinstance(r.get("total"),(int,float)))
    st.markdown(f"""<div class="stats">
      <div class="stat"><span class="stat-val">{len(all_r)}</span><span class="stat-lbl">Receipts</span></div>
      <div class="stat"><span class="stat-val">{total_items}</span><span class="stat-lbl">Line items</span></div>
      <div class="stat"><span class="stat-val">${grand:.2f}</span><span class="stat-lbl">Grand total</span></div>
    </div>""",unsafe_allow_html=True)
 
    # Manual edit panel
    render_edit_panel(all_r)
 
    st.markdown("---")
 
    # Spend by store
    store_totals={}
    for rd in all_r:
        t=rd.get("total")
        if isinstance(t,(int,float)): store_totals[rd["store_name"]]=store_totals.get(rd["store_name"],0)+t
    if store_totals:
        st.markdown('<div class="section-hdr">📊 Spend by store</div>',unsafe_allow_html=True)
        st.markdown('<div class="analysis-card"><div class="bar-wrap">',unsafe_allow_html=True)
        mx=max(store_totals.values())
        for s,a in sorted(store_totals.items(),key=lambda x:-x[1]):
            st.markdown(f'<div class="bar-row"><div class="bar-label" title="{s}">{s}</div><div class="bar-track"><div class="bar-fill" style="width:{a/mx*100:.1f}%"></div></div><div class="bar-value">${a:.2f}</div></div>',unsafe_allow_html=True)
        st.markdown('</div></div>',unsafe_allow_html=True)
 
    # Price comparison
    item_map={}
    for rd in all_r:
        for it in rd["items"]: item_map.setdefault(it["name"].lower().strip(),[]).append({"store":rd["store_name"],"price":it["price"]})
    dupes={k:v for k,v in item_map.items() if len(v)>1}
    if dupes:
        st.markdown('<div class="section-hdr">🏷️ Price comparison</div>',unsafe_allow_html=True)
        st.markdown('<div class="analysis-card">',unsafe_allow_html=True)
        for name,entries in dupes.items():
            bp=min(e["price"] for e in entries); bs=next(e["store"] for e in entries if e["price"]==bp)
            st.markdown(f'<div class="best-row"><div><span class="best-name">{name.title()}</span><span class="best-tag">best deal</span><br><span class="best-store">at {bs}</span></div><span class="best-price">${bp:.2f}</span></div>',unsafe_allow_html=True)
        st.markdown('</div>',unsafe_allow_html=True)
    else:
        st.markdown('<div class="tip">ℹ️ <strong>Price comparison</strong> — scan receipts from multiple stores to compare item prices.</div>',unsafe_allow_html=True)
 
    # Export
    st.markdown('<div class="section-hdr">📥 Export</div>',unsafe_allow_html=True)
    c1,c2=st.columns([3,1])
    with c1:
        xl=build_excel(all_r); ts=datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        st.download_button(label="📥 Download Excel report — Summary · Line Items · Price Comparison",
            data=xl,file_name=f"ReceiptIQ_{ts}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",key="xl_dl")
    with c2:
        if st.button("🗑 Clear all"):
            st.session_state["receipts"]=[]; st.rerun()
 
st.markdown("---")
st.markdown('<div style="display:flex;justify-content:space-between;font-size:.72rem;color:#4e5a78;"><span>ReceiptIQ · FIN4901 FinTech Graduate Project</span><span>YOLO26 · EasyOCR · OpenCV · Streamlit</span></div>',unsafe_allow_html=True)
